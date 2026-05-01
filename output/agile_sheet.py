import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date
import os

HEADERS = ["Date","Sprint","Task / User Story","Status","Priority","Assigned To","Blockers","Notes"]

def init_agile_sheet(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AGILE Tracker"
    ws.append(HEADERS)
    for col, cell in enumerate(ws[1], 1):
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 30
    wb.save(path)
    print(f"  AGILE sheet created: {path}")

def update_agile_sheet(path, task, status, notes, sprint="Sprint 1", priority="Must Have", assigned_to="Auto"):
    if not os.path.exists(path):
        init_agile_sheet(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.append([str(date.today()), sprint, task, status, priority, assigned_to, "", notes])
    wb.save(path)
    print(f"  AGILE updated: {task} → {status}")